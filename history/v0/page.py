import tkinter as tk
from tkinter import font  as tkfont
from tkinter import messagebox
import os
import docx2txt
import openpyxl as oxl
from docxtpl import DocxTemplate
import win32com.client
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
import smtplib

def getData(handle):
    data = []
    for x in handle.rows:
        temp = []
        for y in x:
            temp.append(y.value)
        data.append(temp)
    return data

def createData(data):
    cleanData = []
    for x in range(1,len(data)):
    	temp = {}
    	for y in range(len(data[0])):
    		temp[data[0][y]] = data[x][y]
    	cleanData.append(temp)
    return cleanData

wordFile = ''
def getWord():
    global wordFile
    cert = [f for f in os.listdir('.') if os.path.isfile(f) and (f.split(".")[1]=='docx' or f.split(".")[1]=='doc')][0]
    wordFile = os.getcwd()+'\\'+cert
getWord()

certiFolder = os.getcwd()+"\\certificates\\"

def createDB():
    global wordFile
    data = docx2txt.process(wordFile)
    data = data.split("}}")[:-1]
    data = [x.split("{{")[-1].strip() for x in data]
    data.append("EMAIL_ID")
    wb = oxl.Workbook()
    ws = wb.active
    ws.append(data)
    a = wb.save('database.xlsx')
    os.startfile("database.xlsx")

root = tk.Tk()

def database():
    if(os.path.isfile("database.xlsx")):
        inp = messagebox.askyesno("Already Exists","The File you are trying to create alredy exists! Overwrite?")
        if(inp):
            createDB()
        else:
            os.startfile("database.xlsx")
    else:
        createDB()
    global button1
    button1.configure(bg = "#00ff11")
    button1.configure(text='Create Database '+u'\u2713')

def makeCerti(entry,counter,n):
    try:
        global wordFile
        wordCerti = certiFolder + f"{counter}.docx"
        pdfCerti = certiFolder + f"{counter}.pdf"
        document = DocxTemplate(wordFile)
        document.render(entry)
        document.save(wordCerti)
        word = win32com.client.Dispatch('Word.Application')
        doc = word.Documents.Open(wordCerti)
        doc.SaveAs(pdfCerti, FileFormat=17)
        doc.Close()
        word.Quit()
        aob.configure(text=f"{counter}/{n}")
        root.update()
    except Exception as e:
        print(e)
        input()
        makeCerti(entry,counter,n)

def create():
    global button2
    button2.grid_forget()
    button2 = tk.Button(root, text='Create Certificates', font=buttonFont, padx=5, pady=2, command=create)
    button2.grid(row=2, column=0, padx=50,pady=15)
    global fileLocation
    wb = oxl.load_workbook(os.getcwd()+'\\database.xlsx')
    ws = wb.active
    data = getData(ws)
    cleanData = createData(data)
    if(cleanData):
        button2.grid_forget()
        button2 = tk.Button(root, text='Please Wait ..', font=buttonFont, padx=5, pady=2, command=create)
        button2.grid(row=2, column=0, padx=50,pady=15)
        aob.configure(text=f"0/{len(cleanData)}")
        root.update()
        counter = 1
        os.makedirs(os.getcwd()+"/certificates", exist_ok=True)
        for entry in cleanData:
            print(entry)
            makeCerti(entry,counter,len(cleanData))
            counter+=1
        button2.grid_forget()
        button2 = tk.Button(root, text='Create Certificates '+u'\u2713', bg = "#00ff11", font=buttonFont, padx=5, pady=2, command=create)
        button2.grid(row=2, column=0, padx=50,pady=15)
        aob.configure(text=f"{counter-1}/{len(cleanData)}")
    else:
        messagebox.showinfo("No Data","Please fill the database file to generate certificates.")

def createAndSend(send_to, subject, message, certiFile):
    send_from = "deekshantwadhwa2000@gmail.com"
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Subject'] = subject
    msg.attach(MIMEText(message))
    part = MIMEBase('application', "octet-stream")
    with open(certiFile, 'rb') as file:
        part.set_payload(file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition','attachment; filename="{}"'.format(Path(certiFile).name))
    msg.attach(part)
    smtp = smtplib.SMTP(host="smtp.gmail.com", port= 587)
    smtp.starttls()
    smtp.login(send_from, "ldqsjwdxgkzqimsv")
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()

def email():
    global certiFolder
    global button3
    button3.grid_forget()
    button3 = tk.Button(root, text='Email Certificates', font=buttonFont, padx=5, pady=2, command=email)
    button3.grid(row=3, column=0, padx=50,pady=(15,60))
    pdfs = [f for f in os.listdir(certiFolder) if f.split(".")[-1]=='pdf']
    wb = oxl.load_workbook(os.getcwd()+'\\database.xlsx')
    ws = wb.active
    data = getData(ws)
    ids = []
    if('EMAIL_ID' in data[0]):
        emailIndex = data[0].index('EMAIL_ID')
        for x in range(1,len(data)):
            ids.append(data[x][emailIndex])
        counter = 1
        button3.grid_forget()
        button3 = tk.Button(root, text='Please Wait ..', font=buttonFont, padx=5, pady=2, command=create)
        button3.grid(row=3, column=0, padx=50,pady=(15,60))
        eob.configure(text=f"0/{len(ids)}")
        root.update()
        for x in range(len(ids)):
            createAndSend(ids[x], "Your Certificate", "Thank you for taking part in this contest", certiFolder+pdfs[x])
            eob.configure(text=f"{counter}/{len(ids)}")
            root.update()
            counter+=1
        button3.grid_forget()
        button3 = tk.Button(root, text='Email Certificates '+u'\u2713', bg = "#00ff11", font=buttonFont, padx=5, pady=2, command=create)
        button3.grid(row=3, column=0, padx=50,pady=(15,60))
        eob.configure(text=f"{counter-1}/{len(ids)}")
    else:
        messagebox.showinfo("No Emails Found!","PLease fill the EMAIL_ID field in the database.")

def test():
    root.destroy()

tk.Label(root, text='Certificate Maker', font=("Verdana", 25)).grid(row=0, column=0, padx=60,pady=(10,30))
buttonFont = tkfont.Font(size=15)
button1 = tk.Button(root, text='Create Database', font=buttonFont, padx=5, pady=2, command=database)
button1.grid(row=1, column=0, padx=50,pady=15)
button2 = tk.Button(root, text='Create Certificates', font=buttonFont, padx=5, pady=2, command=create)
button2.grid(row=2, column=0, padx=50,pady=15)
button3 = tk.Button(root, text='Email Certificates', font=buttonFont, padx=5, pady=2, command=email)
button3.grid(row=3, column=0, padx=50,pady=(15,60))
button4 = tk.Button(root, text='Exit', font=buttonFont, padx=5, pady=2, command=test)
button4.grid(row=4, column=0, padx=50,pady=(15,60))
aob = tk.Label(root,text='')
aob.grid(row=2,column=0,padx=(250,0))
qweasd = tk.Label(root,text='')
qweasd.grid(row=2,column=0,padx=(250,0),pady=60)
eob = tk.Label(root,text='')
eob.grid(row=3,column=0,padx=(250,0))
root.mainloop()
