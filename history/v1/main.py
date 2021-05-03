import tkinter as tk
from tkinter import font  as tkfont
from tkinter import filedialog
from tkinter import messagebox

import os
from shutil import copyfile, rmtree
import json

import docx2txt
import openpyxl as oxl
from docxtpl import DocxTemplate

import win32com.client

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import base64

import datetime

tick = u'\u2713'

def create_project(controller,filename):
    folderName = filename.split("/")[-1].split(".")[0].strip().replace(" ","-")
    docType = filename.split(".")[-1]
    fnf['projectName']          = folderName
    fnf['projectFolder']        = fnf['projects']+f"/{folderName}"
    fnf['projectTemplate']      = fnf['projectFolder'] + f"/template.{docType}"
    fnf['projectSettings']      = fnf['projectFolder'] + "/settings"
    fnf['projectCertificates']  = fnf['projectFolder'] + "/Certificates"
    fnf['projectDatabase']      = fnf['projectFolder'] + "/database.xlsx"
    if(not os.path.isdir(fnf['projects'])):
        os.mkdir(fnf['projects'])
    os.mkdir(fnf['projectFolder'])
    copyfile(filename,fnf['projectTemplate'])
    createSettingsFile(fnf['projectSettings'])
    controller.show_frame("Project")

def new_project(controller):
    filename = filedialog.askopenfilename(initialdir = ".",title = 'Select Certificate Template', filetypes = (("Word Template",("*.doc","*.docx")),))
    if(filename):
        folderName = filename.split("/")[-1].split(".")[0].strip().replace(" ","-")
        if(os.path.isdir(fnf['projects']+f"\\{folderName}")):
            inp = messagebox.askyesno("Already Exists","Project with this name already exists! Overwrite?")
            if(inp):
                rmtree(fnf['projects']+f"\\{folderName}")
                create_project(controller,filename)
        else:
            create_project(controller,filename)

def existing_project(controller):
    filename = filedialog.askdirectory(initialdir=fnf['projects'])
    fnf['projectName']          = filename.split("/")[-1].strip()
    fnf['projectFolder']        = filename
    fnf['projectTemplate']      = fnf['projectFolder'] + "/" + [f for f in os.listdir(fnf['projectFolder']) if (f=='template.doc' or f=='template.docx' or f=='template.dox')][0]
    fnf['projectSettings']      = fnf['projectFolder'] + "/settings"
    fnf['projectCertificates']  = fnf['projectFolder'] + "/Certificates"
    fnf['projectDatabase']      = fnf['projectFolder'] + "/database.xlsx"
    if(not os.path.exists(filename)):
        messagebox.showerror("Project Doesnot Exists","The project you have chosen doesnot exists!")
        return
    if(not os.path.exists(fnf['projectSettings'])):
        createSettingsFile(fnf['projectSettings'])
    if(not os.path.exists(fnf['projectTemplate'])):
        inp = messagebox.askyesno("Template dosen't exist!","The project you have chosen doesnot contain a template file! Continue?")
        if(inp):
            controller.show_frame("Project")
    else:
        controller.show_frame("Project")

def createDB():
    data = docx2txt.process(fnf['projectTemplate'])
    data = data.split("}}")[:-1]
    data = [x.split("{{")[-1].strip() for x in data]
    data.append("EMAIL_ID")
    wb = oxl.Workbook()
    ws = wb.active
    ws.append(data)
    a = wb.save(fnf['projectDatabase'])
    os.startfile(fnf['projectDatabase'])

def database(thisButton):
    if(os.path.isfile(fnf['projectDatabase'])):
        inp = messagebox.askyesno("Already Exists","Database for this project alredy exists! Overwrite?")
        if(inp):
            createDB()
        else:
            os.startfile(fnf['projectDatabase'])
    else:
        createDB()
    thisButton.configure(bg = "#00ff11")
    thisButton.configure(text='Create Database')

def getData(handle):
    data = []
    for x in handle.rows:
        temp = []
        for y in x:
            temp.append(y.value)
        data.append(temp)
    return data

def createData(data):
    cleanData = []
    for x in range(1,len(data)):
    	temp = {}
    	for y in range(len(data[0])):
            if(data[0][y] and data[x][y]):
                temp[data[0][y]] = data[x][y]
    	cleanData.append(temp)
    return cleanData

def makeCerti(counter):
    wordCerti = fnf['projectCertificates'] + f"/{counter}.docx" 
    pdfCerti = fnf['projectCertificates'] + f"/{counter}.pdf"
    try:
        doc = wordHandle.Documents.Open(wordCerti)
        # wordHandle.Visible = True
        doc.SaveAs(pdfCerti, FileFormat=17)
        doc.Close()
    except Exception as e:
        print(wordCerti)
        print(e)
        if(input()=='exit'):
            return
        makeCerti(counter)

def create(thisButton,controller):
    thisButton.configure(background='SystemButtonFace')
    thisButton.configure(text='Create Certificates')
    if(not os.path.isfile(fnf['projectDatabase'])):
        createDB()
    wb = oxl.load_workbook(fnf['projectDatabase'])
    ws = wb.active
    data = getData(ws)
    cleanData = createData(data)
    if(cleanData):
        counter = 1
        if(os.path.exists(fnf['projectCertificates'])):
            inp = messagebox.askyesno("Already Exists","You might already have some certificates in your Project directory. Overwrite?")
            if(inp):
                try:
                    a = rmtree(fnf['projectCertificates'])
                    while(os.path.exists(fnf['projectCertificates'])):
                        pass
                except:
                    #use google api here
                    pass
            else:
                return
        os.makedirs(fnf['projectCertificates'], exist_ok=True)
        thisButton.configure(text='Please Wait ..')
        aob.configure(text="6 sec")
        controller.update()
        global wordHandle
        # wordHandle = win32com.client.Dispatch('Word.Application')
        wordHandle = win32com.client.DispatchEx('Word.Application')
        for entry in cleanData:
            wordCerti = fnf['projectCertificates'] + f"/{counter}.docx"
            document = DocxTemplate(fnf['projectTemplate'])
            document.render(entry)
            document.save(wordCerti)
            makeCerti(counter)
            aob.configure(text=f"{counter}/{len(cleanData)}")
            controller.update()
            counter+=1
        wordHandle.Quit()
        if(not os.path.exists(fnf['baseDatabase'])):
            create_records_file()
        wb = oxl.load_workbook(fnf['baseDatabase'])
        ws = wb.active
        ws.append(["Project : ",fnf['projectName']])
        ws.append(["Created : ",str(datetime.datetime.now().date()),str(datetime.datetime.now().time())])
        ws.append([" "," "])
        emptyLineB4First = 0
        for x in data:
            ws.append(x)
            if(not emptyLineB4First):
                ws.append([" "," "])
                emptyLineB4First=1
        ws.append([" "," "])
        ws.append(["_","_","_","_","_","_","_",])
        ws.append([" "," "])
        ws.append([" "," "])
        wb.save(fnf['baseDatabase'])
        thisButton.configure(text='Create Certificates')
        thisButton.configure(bg = "#00ff11")
        aob.configure(text=f"Done({len(cleanData)})!")
    else:
        messagebox.showinfo("No Data","Please fill the database file to generate certificates.")

def createSettingsFile(l):
    handle = open(l,'w')
    settings = {
        "emailSubject"  :"Here is you Certificate!",
        "emailBody"     :"Thank You for your participation.",
    }
    handle.write(json.dumps(settings))
    handle.close()

def createAndSend(send_from,send_to,certiFile,message,smtp,subject):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['Subject'] = subject
    msg['To'] = send_to
    msg.attach(MIMEText(message))
    part = MIMEBase('application', "octet-stream")
    with open(certiFile, 'rb') as file:
        part.set_payload(file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition','attachment; filename="Offer Letter.pdf"')
    msg.attach(part)
    # use async or multithreading
    smtp.sendmail(send_from, send_to, msg.as_string())

def email(thisButton,controller):
    wb = oxl.load_workbook(fnf['projectDatabase'])
    ws = wb.active
    data = getData(ws)
    if(not os.path.exists(fnf['projectCertificates'])):
        messagebox.showinfo("No Data","Please create the certificates first in order to Email them.")
        return
    if('EMAIL_ID' in data[0]):
        thisButton.configure(background='SystemButtonFace')
        thisButton.configure(text='Email Certificates')
        pdfs = [f for f in os.listdir(fnf['projectCertificates']) if f.split(".")[-1]=='pdf']
        if(not len(pdfs)):
            messagebox.showinfo("No Data","Please create the certificates first in order to Email them.")
            return
        emailIndex = data[0].index('EMAIL_ID')
        ids = []
        for x in range(1,len(data)):
            ids.append(data[x][emailIndex])
        counter = 1
        thisButton.configure(text='Please Wait ..')
        eob.configure(text="8 sec")
        controller.update()
        if(not os.path.exists(fnf['projectSettings'])):
            createSettingsFile(fnf['projectSettings'])
        settings = json.loads(open(fnf['projectSettings']).read())
        baseSettings = json.loads((base64.b64decode(base64.b64decode(base64.b64decode(base64.b64decode(base64.b64decode(open(fnf['baseSettings'],'rb').read())))))).decode('utf8').replace("'", '"'))
        send_from = baseSettings['emailFrom']
        subject = settings['emailSubject']
        message = settings['emailBody']
        inputData = baseSettings['inputData']
        smtp = smtplib.SMTP(host="smtp.gmail.com", port= 587)
        smtp.starttls()
        smtp.login(send_from, inputData)
        for x in range(len(ids)):
            pdfLocation = fnf['projectCertificates']+"/"+pdfs[x]
            createAndSend(send_from,ids[x],pdfLocation,message,smtp,subject)
            eob.configure(text=f"{counter}/{len(ids)}")
            controller.update()
            counter+=1
        smtp.quit()
        thisButton.configure(text='Email Certificates')
        thisButton.configure(bg = "#00ff11")
        eob.configure(text=f"{counter-1}/{len(ids)}")
        eob.configure(text=f"Done({len(ids)})!")
    else:
        messagebox.showinfo("No Emails Found!","PLease fill the EMAIL_ID field in the database.")

##########################################################
#########################################################
#########################################################

def project_settings(controller):
    controller.show_frame("projectSettings")
    settingsHandle = json.loads(open(fnf['projectSettings'],'r').read())
    emailSubjectText.delete('1.0', "end-1c")
    emailBodyText.delete('1.0', "end-1c")
    emailSubjectText.insert("end-1c", settingsHandle['emailSubject'])
    emailBodyText.insert("end-1c", settingsHandle['emailBody'])

def project_settings_save(controller):
    subjectText = emailSubjectText.get("1.0","end-1c")
    emailSubjectText.delete('1.0', "end-1c")
    bodyText = emailBodyText.get("1.0","end-1c")
    emailBodyText.delete('1.0', "end-1c")
    settingsHandle = json.loads(open(fnf['projectSettings'],'r').read())
    settingsHandle['emailSubject'] = subjectText
    settingsHandle['emailBody'] = bodyText
    handle = open(fnf['projectSettings'],'w')
    handle.write(json.dumps(settingsHandle))
    handle.close()
    controller.show_frame("Project")

def create_records_file():
    # print(datetime.datetime.now().date())
    # print(datetime.datetime.now().time())
    # wb = oxl.load_workbook(fnf['baseDatabase'])

    wb = oxl.Workbook()
    ws = wb.active
    data = [["Records File"],["Created on:",str(datetime.datetime.now().date())],["Created At:",str(datetime.datetime.now().time())],["",""],["_","_","_","_"],[" ","  "],["  "," "],[" "," "]]
    for x in data:
        ws.append(x)
    wb.save(fnf['baseDatabase'])

def erase_and_create_main_database(thisButton):
    if(os.path.exists(fnf['baseDatabase'])):
        inp = messagebox.askyesno("Already Exists","You might already have some certificates in your Project directory. Overwrite?")
        if(inp):
            create_records_file()
            thisButton.configure(bg = "#00ff11")
    else:
        create_records_file()
        thisButton.configure(bg = "#00ff11")

def test(controller):
    create_records_file()
    print(2)
    # wb = oxl.load_workbook(fnf['projectDatabase'])
    # ws = wb.active
    # wb = oxl.load_workbook("C:\\Users\\Deekshant\\Desktop\\certi\\Certificate-Maker\\records.xlsx")
    # data = [1,2,3]
    # ws = wb.active
    # data = getData(ws)
    # print(data)
    # ws.append(data)
    # wb.save("C:\\Users\\Deekshant\\Desktop\\certi\\Certificate-Maker\\records.xlsx")

class MainApp(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        # self.geometry("500x600")
        # self.resizable(0,0)
        self.title_font = tkfont.Font(family='Verdana', size=25, underline=1)
        self.buttonFont = tkfont.Font(size=15)
        self.settingsFont = tkfont.Font(family='Verdana', size=18, underline=1)
        container = tk.Frame(self)
        container.grid(row=0, column=0)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.frames = {}
        for F in (StartPage, PageOne, Project, mainSettings, projectSettings):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame("StartPage")
    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()

class StartPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60,pady=(10,25))
        button1 = tk.Button(self, text='New Project', font=controller.buttonFont, padx=5, pady=2, command=lambda: new_project(controller))
        button1.grid(row=1, column=0, padx=50,pady=30)
        button2 = tk.Button(self, text='Existing Project', font=controller.buttonFont, padx=5, pady=2, command=lambda: existing_project(controller))
        button2.grid(row=2, column=0, padx=50,pady=30)
        button3 = tk.Button(self, text='Database', font=controller.buttonFont, padx=5, pady=2)
        button3.grid(row=3, column=0, padx=50,pady=30)
        button4 = tk.Button(self, text='Settings', font=controller.buttonFont, padx=5, pady=2, command=lambda: controller.show_frame("mainSettings"))
        button4.grid(row=4, column=0, padx=50,pady=30)
        # button5 = tk.Button(self, text='Test', font=controller.buttonFont, padx=5, pady=2, command=lambda: test(controller))
        # button5.grid(row=5, column=0, padx=50,pady=(30,60))

class Project(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60,pady=(10,25))
        button1 = tk.Button(self, text='Create Database', font=controller.buttonFont, padx=5, pady=2, command=lambda: database(button1))
        button1.grid(row=1, column=0, padx=50,pady=30)
        button2 = tk.Button(self, text='Create Certificates', font=controller.buttonFont, padx=5, pady=2, command=lambda: create(button2,controller))
        button2.grid(row=2, column=0, padx=50,pady=30)
        button3 = tk.Button(self, text='Email Certificates', font=controller.buttonFont, padx=5, pady=2, command=lambda: email(button3,controller))
        button3.grid(row=3, column=0, padx=50,pady=30)
        button4 = tk.Button(self, text='Settings', font=controller.buttonFont, padx=5, pady=2, command=lambda: project_settings(controller))
        button4.grid(row=4, column=0, padx=50,pady=(30,60))
        global aob
        aob = tk.Label(self,text='')
        aob.grid(row=2,column=0,padx=(250,0))
        global eob
        eob = tk.Label(self,text='')
        eob.grid(row=3,column=0,padx=(250,0))

class mainSettings(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60,pady=(10,25))
        sLabel = tk.Label(self, text="Main Settings", font=controller.settingsFont)
        sLabel.grid(row=1,column=0)
        button1 = tk.Button(self, text='Delete Records', font=controller.buttonFont, padx=5, pady=2, command=lambda: erase_and_create_main_database(button1))
        button1.grid(row=2, column=0, padx=50,pady=30)
        # button2 = tk.Button(self, text='Create Certificates', font=controller.buttonFont, padx=5, pady=2, command=lambda: create(button2,controller))
        # button2.grid(row=2, column=0, padx=50,pady=30)
        # button3 = tk.Button(self, text='Email Certificates', font=controller.buttonFont, padx=5, pady=2, command=lambda: email(button3,controller))
        # button3.grid(row=3, column=0, padx=50,pady=30)
        button4 = tk.Button(self, text='Back', font=controller.buttonFont, padx=5, pady=2, command=lambda: controller.show_frame("StartPage"))
        button4.grid(row=4, column=0, padx=50,pady=(30,60))

class projectSettings(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60,pady=(10,15))
        sLabel = tk.Label(self, text="Project Settings", font=controller.settingsFont)
        sLabel.grid(row=1,column=0, padx=50,pady=(0,10))
        emailSubject = tk.Label(self, text="Email Subject :", font=("Verdana", 12))
        emailSubject.grid(row=2,column=0,sticky="W",padx=30,pady=(10,5))

        global emailSubjectText
        emailSubjectText = tk.Text(self, bd=1, width=0, height=3, font=("Verdana", 10))
        emailSubjectText.grid(row=3,column=0,sticky='we', padx=30)

        emailBody = tk.Label(self, text="Email Body :", font=("Verdana", 12))
        emailBody.grid(row=4,column=0,sticky="W",padx=30,pady=(30,5))

        global emailBodyText
        emailBodyText = tk.Text(self, bd=1, width=0, height=5, font=("Verdana", 10))
        emailBodyText.grid(row=5,column=0,sticky='we', padx=30)

        button2 = tk.Button(self, text='Save Changes', font=controller.buttonFont, padx=5, pady=2, command=lambda: project_settings_save(controller))
        button2.grid(row=7, column=0, padx=50,pady=(20,10))
        # button3 = tk.Button(self, text='Email Certificates', font=controller.buttonFont, padx=5, pady=2, command=lambda: email(button3,controller))
        # button3.grid(row=3, column=0, padx=50,pady=30)
        button4 = tk.Button(self, text='Back', font=controller.buttonFont, padx=5, pady=2, command=lambda: controller.show_frame("Project"))
        button4.grid(row=8, column=0, padx=50,pady=(20,0))


class PageOne(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)
        button = tk.Button(self, text="Go to the start page",command=lambda: controller.show_frame("StartPage"))
        button.pack()

if __name__ == "__main__":
    fnf = {}
    fnf['current'] = os.getcwd()
    fnf['projects'] = fnf['current'] + "\\projects"
    fnf['baseSettings'] = fnf['current'] + "\\input.dk"
    fnf['baseDatabase'] = fnf['current'] + "\\Records.xlsx"
    app = MainApp()
    app.mainloop()
