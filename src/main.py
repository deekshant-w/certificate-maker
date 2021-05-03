# gui components
import tkinter as tk
from tkinter import font as tkfont
from tkinter import filedialog
from tkinter import messagebox

# data and file management
import os
from shutil import copyfile, rmtree
import json

# working with word file
# import docx2txt
import openpyxl as oxl
from docxtpl import DocxTemplate

# word to pdf
import win32com.client
from pathlib import Path

# create and send email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import re

# date and time handling
import datetime
from time import sleep, time

# randomization
from string import ascii_uppercase, digits
import random


# green tick unicode
tick = "\u2713"

# charecters for random name generation
chars = ascii_uppercase + digits


def makeNewName(size=6):
    """
    Generates random names
    """
    res = (
        "".join([random.choice(ascii_uppercase) for _ in range(2)])
        + str(random.randint(10, 99))
        + "".join([random.choice(chars) for _ in range(size)])
        + str(int(time()) % (10 ** size))
    )
    return res


def createCredsFile(location):
    """
    Create credentials file if doesnot exists
    """
    if os.path.isfile(location):
        return

    credText = "email - \npassword- "
    file = open(location, "w")
    file.write(credText)
    file.close()


def create_project(controller, filename):
    """
    Initialize new project folder with required files
    """

    # new folder name for project
    folderName = filename.split("/")[-1].split(".")[0].strip().replace(" ", "-")

    # template file extension (doc/docx)
    docType = filename.split(".")[-1]

    fnf["projectName"] = folderName
    fnf["projectFolder"] = fnf["projects"] + f"/{folderName}"
    fnf["projectTemplate"] = fnf["projectFolder"] + f"/template.{docType}"

    # settings file for current project
    fnf["projectSettings"] = fnf["projectFolder"] + "/settings"

    fnf["projectCertificates"] = fnf["projectFolder"] + "/Certificates"
    fnf["projectDatabase"] = fnf["projectFolder"] + "/database.xlsx"

    # if projects folder is not created, create it
    if not os.path.isdir(fnf["projects"]):
        os.mkdir(fnf["projects"])

    # create folder for current project
    os.mkdir(fnf["projectFolder"])

    # copy template file to new project folder
    copyfile(filename, fnf["projectTemplate"])

    # create settings file
    createSettingsFile(fnf["projectSettings"], template=f"template.{docType}")

    # switch to project page view
    controller.show_frame("Project")


def new_project(controller):
    """
    Create new project using the the input word 
    template file. The project is started with project
    name, same as the word file.
    """

    # prompt for template location
    filename = filedialog.askopenfilename(
        initialdir=".",
        title="Select Certificate Template",
        filetypes=(("Word Template", ("*.doc", "*.docx")),),
    )

    # if a word file is passed and is valid
    if filename:

        # new folder name
        folderName = filename.split("/")[-1].split(".")[0].strip().replace(" ", "-")

        # if folder with that name already exists
        if os.path.isdir(fnf["projects"] + f"\\{folderName}"):
            inp = messagebox.askyesno(
                "Already Exists", "Project with this name already exists! Overwrite?"
            )

            # Overwrite previous project
            if inp:
                # delete old project
                rmtree(fnf["projects"] + f"\\{folderName}")
                # start new project initialzations
                create_project(controller, filename)
        else:
            # start new project initialzations
            create_project(controller, filename)


def existing_project(controller):
    """
    Loads a pre existing project
    (template file + settings file)
    """

    # prompt for project folder
    filename = filedialog.askdirectory(initialdir=fnf["projects"])

    if not os.path.exists(filename):
        messagebox.showerror(
            "Project Doesnot Exists", "The project you have chosen doesnot exists!"
        )
        return

    # project details
    fnf["projectName"] = filename.split("/")[-1].strip()
    fnf["projectFolder"] = filename
    fnf["projectSettings"] = fnf["projectFolder"] + "/settings"

    if not os.path.exists(fnf["projectSettings"]):
        createSettingsFile(fnf["projectSettings"])

    fnf["projectCertificates"] = fnf["projectFolder"] + "/Certificates"
    fnf["projectDatabase"] = fnf["projectFolder"] + "/database.xlsx"

    # load setting file
    try:
        settingsFile = json.loads(open(fnf["projectSettings"]).read())
    except:
        settingsFile = {}

    # retrieve word file if template doesnot exist
    try:
        fileTemplate = settingsFile["templateFile"]
    except:
        fileTemplate = ""
        for f in os.listdir(fnf["projectFolder"]):
            if f == "template.doc" or f == "template.docx" or f == "template.dox":
                fileTemplate = f
                settingsFile["templateFile"] = f
                open(fnf["projectSettings"], "w").write(json.dumps(settingsFile))
                break

    fnf["projectTemplate"] = fnf["projectFolder"] + "/" + fileTemplate

    if not os.path.exists(fnf["projectTemplate"]):
        inp = messagebox.askyesno(
            "Template dosen't exist!",
            "The project you have chosen doesnot contain a template file! Continue?",
        )
        if inp:
            controller.show_frame("Project")
    else:
        controller.show_frame("Project")


def createDB():
    # # Extract placeholders from template file
    # data = docx2txt.process(fnf["projectTemplate"])
    # data = data.split("}}")[:-1]
    # data = [x.split("{{")[-1].strip() for x in data]

    # Extract placeholders using jinjs
    data = list(
        DocxTemplate(fnf["projectTemplate"]).get_undeclared_template_variables()
    )
    data = ["EMAIL_ID", "CERTIFICATE_CREATED", "MAIL_SENT"] + data

    # creating workbook with "data" columns
    wb = oxl.Workbook()
    ws = wb.active
    ws.append(data)

    wb.save(fnf["projectDatabase"])
    os.startfile(fnf["projectDatabase"])


def database(thisButton):
    """
    If datadase exists then open it,
    else create it and open it
    """
    if os.path.isfile(fnf["projectDatabase"]):
        os.startfile(fnf["projectDatabase"])
    else:
        createDB()

    ## Make the button gren and change its text
    # thisButton.configure(bg="#00ff11")
    # thisButton.configure(text="Create Database")


def getData(handle):
    """
    extract all data from excel sheet
    row-wise.
    row[0] -> column headers
    """
    data = []
    for x in handle.rows:
        # rows
        temp = []
        for y in x:
            temp.append(y.value)
        data.append(temp)
    return data


def createData(data):
    """
    Convert [[],[]] data into [{},{}]
    for to be used in jinja.
    data[0] -> column headers
    """
    cleanData = []

    # skipping header row
    for x in range(1, len(data)):
        temp = {}
        for y in range(len(data[0])):
            # if column header exists and cell value exists
            if data[0][y] and data[x][y]:
                temp[data[0][y].strip()] = data[x][y]
        cleanData.append(temp)
    return cleanData


# maximum number of retries before pdf conversion fails
MAX_RETRIES = 3


def makeCerti(filename, retry=MAX_RETRIES):
    if not retry:
        return 0
    wordCerti = os.fspath(Path(fnf["projectCertificates"] + f"/{filename}.docx"))
    pdfCerti = os.fspath(Path(fnf["projectCertificates"] + f"/{filename}.pdf"))
    try:
        doc = wordHandle.Documents.Open(wordCerti)
        # wordHandle.Visible = True
        doc.SaveAs(pdfCerti, FileFormat=17)
        doc.Close()
        return 1
    except Exception as e:
        print()
        print(e.args)
        print(e.excepinfo)
        print(e.hresult)
        sleep(1 / (1 + MAX_RETRIES - retry))
        makeCerti(filename, retry - 1)


def create(thisButton, controller):
    """
    Create certificates using database
    in certificates folder of project
    """

    # normal configuration of button
    thisButton.configure(background="SystemButtonFace")
    thisButton.configure(text="Create Certificates")
    thisButton.configure(state="normal")

    if not os.path.isfile(fnf["projectDatabase"]):
        createDB()

    wb = oxl.load_workbook(fnf["projectDatabase"])
    ws = wb.active

    data = getData(ws)
    cleanData = createData(data)

    # if sheet has any data except headers
    if cleanData:
        # if certificates folder exists then
        # there might be some ceertificates in it
        if os.path.exists(fnf["projectCertificates"]):
            inp = messagebox.askyesno(
                "Already Exists",
                "You might already have some certificates in your Project directory. Overwrite?",
            )

            if inp:
                try:
                    # deleting certificates folder
                    rmtree(fnf["projectCertificates"])

                    # waiting till certificates folder exists
                    while os.path.exists(fnf["projectCertificates"]):
                        pass
                except:
                    pass
            else:
                return

        # create certificates folder
        os.makedirs(fnf["projectCertificates"], exist_ok=True)

        # update "Create certificates button"
        thisButton.configure(text="Please Wait ..")
        thisButton.configure(state="disable")

        # update waiting label
        # aob.configure(text="6 sec")

        # manually update GUI
        controller.update()

        # word handle to convert word -> pdf on windows
        # shared across all functions and all calls
        global wordHandle

        # Dispatch -> creates new instance
        # wordHandle = win32com.client.Dispatch('Word.Application')

        # DispatchEx -> searches for word instance in background
        wordHandle = win32com.client.DispatchEx("Word.Application")

        # hide word instance, work in background
        wordHandle.Visible = False

        # counter for number of entries in database
        counter = 0

        fileList = []

        for entry in cleanData:
            # updating counter
            aob.configure(text=f"{counter}/{len(cleanData)}")
            controller.update()
            counter += 1

            # getting CERTIFICATE_CREATED for that row
            # dont create if already created
            filename = entry.get("CERTIFICATE_CREATED", "")
            if filename.lower() not in ["no", "0", "-", ""]:
                fileList.append(filename)
                continue

            # if certificate is not already created
            # create new filename
            filename = makeNewName()

            # create word certificate
            wordCerti = fnf["projectCertificates"] + f"/{filename}.docx"
            document = DocxTemplate(fnf["projectTemplate"])
            document.render(entry)
            document.save(wordCerti)

            # convert to pdf
            if makeCerti(filename):
                fileList.append(filename)

        # exit word handler
        wordHandle.Quit()

        # saving records
        aob.configure(text=f"Saving!")
        controller.update()

        # save new filenames
        for x in range(len(fileList)):
            row = x + 2
            ws[f"B{row}"] = fileList[x]
        wb.save(fnf["projectDatabase"])

        # create records file
        if not os.path.exists(fnf["baseDatabase"]):
            create_records_file()

        # load records file
        wb = oxl.load_workbook(fnf["baseDatabase"])
        ws = wb.active

        ws.append(["Project : ", fnf["projectName"]])
        ws.append(
            [
                "Created : ",
                str(datetime.datetime.now().date()),
                str(datetime.datetime.now().time()),
            ]
        )
        ws.append([" ", " "])

        # record certificate data
        emptyLineB4First = 0
        for x in data:
            ws.append(x)
            if not emptyLineB4First:
                ws.append([" ", " "])
                emptyLineB4First = 1

        ws.append([" ", " "])
        ws.append(
            ["_", "_", "_", "_", "_", "_", "_",]
        )
        ws.append([" ", " "])
        ws.append([" ", " "])
        wb.save(fnf["baseDatabase"])

        # update GUI
        thisButton.configure(text="Create Certificates")
        thisButton.configure(bg="#00ff11")
        thisButton.configure(state="normal")
        aob.configure(text=f"Done({len(cleanData)})!")

    else:
        messagebox.showinfo(
            "No Data", "Please fill the database file to generate certificates."
        )


def createSettingsFile(l, template=None):
    """
    Create settings file (json) inside 
    new project folder
    """
    handle = open(l, "w")
    settings = {
        "emailSubject": "Here is you Certificate!",
        "emailBody": "Thank You for your participation.",
        "templateFile": template or "",
    }
    handle.write(json.dumps(settings))
    handle.close()


def createAndSend(send_from, send_to, certiFile, message, smtp, subject):
    """
    Create email, Attatch pdf file and send email
    Arguments:
        send_from   - our email
        send_to     - recienver email
        certiFile   - location of pdf file
        message     - message body
        smtp        - smtp object to send email (preset)
        subject     - email subject

    Returns:
        success     - 1
        failure     - 0
    """
    try:
        msg = MIMEMultipart()
        msg["From"] = send_from
        msg["Subject"] = subject
        msg["To"] = send_to
        msg.attach(MIMEText(message))
        part = MIMEBase("application", "octet-stream")
        with open(certiFile, "rb") as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition", 'attachment; filename="Offer Letter.pdf"'
        )
        msg.attach(part)
        # use async or multithreading
        smtp.sendmail(send_from, send_to, msg.as_string())
        return 1
    except Exception as e:
        print(e)
        return 0


def email(thisButton, controller):
    """
    Create, send an record emails
    * Using SMTP
    """
    thisButton.configure(background="SystemButtonFace")
    thisButton.configure(text="Email Certificates")

    # certificate folder doesnot exist
    if not os.path.exists(fnf["projectCertificates"]):
        messagebox.showinfo(
            "No Data", "Please create the certificates first in order to Email them."
        )
        return

    # Email credentials
    createCredsFile(fnf["mailCreds"])
    credFile = open(fnf["mailCreds"]).read().strip().split("\n")

    # sender email
    email = credFile[0]
    email = re.findall(
        r"^[E,e]?[M,m]?[A,a]?[I,i]?[L,l]? *[-,:]* *(.*) *", email
    )[0].strip()

    # sender password
    pas = credFile[1]
    pas = re.findall(
        r"^[P,p]?[A,a]?[S,s]?[S,s]?[W,w]?[O,o]?[R,r]?[D,d]? *[-,:]* *(.*) *", pas
    )[0].strip()

    if not (email and pas):
        messagebox.showinfo("No Credentials Found!", "Please fill the creds.txt file")
        return

    # search if any pdf exists
    pdfs = [
        f for f in os.listdir(fnf["projectCertificates"]) if f.split(".")[-1] == "pdf"
    ]
    if not len(pdfs):
        messagebox.showinfo(
            "No Data", "Please create the certificates first in order to Email them.",
        )
        return

    # load project database
    wb = oxl.load_workbook(fnf["projectDatabase"])
    ws = wb.active
    data = getData(ws)

    # check if email sending field is in database
    if not "EMAIL_ID" in data[0]:
        messagebox.showinfo(
            "No Emails Found!", "PLease fill the EMAIL_ID field in the database."
        )
        return

    # get data for sending email
    data = createData(data)

    # disable email button
    thisButton.configure(text="Please Wait ..")
    thisButton.configure(state="disable")
    controller.update()

    if not os.path.exists(fnf["projectSettings"]):
        createSettingsFile(fnf["projectSettings"])

    # project settings
    settings = json.loads(open(fnf["projectSettings"]).read())

    # email info
    subject = settings["emailSubject"]
    message = settings["emailBody"]

    # setting SMTP server
    smtp = smtplib.SMTP(host="smtp.gmail.com", port=587)
    smtp.starttls()
    smtp.login(email, pas)

    # counter for user display
    counter = 0

    # list for maintaining record in databse
    mailList = []

    for x in data:
        eob.configure(text=f"{counter}/{len(data)}")
        controller.update()
        counter += 1

        # loaction of pdf file for current row
        pdfLocation = (
            fnf["projectCertificates"] + f"/{data['CERTIFICATE_CREATED'] or ''}"
        )
        if os.path.isfile(pdfLocation) and data.get("MAIL_SENT", "").lower() not in [
            "y",
            "yes",
            "1",
        ]:
            # send individual emails
            s = createAndSend(email, x["EMAIL_ID"], pdfLocation, message, smtp, subject)
            if not s:
                # email failed
                mailList.append("failed")
            else:
                # sucessful email
                mailList.append("yes")
        else:
            # email already sent
            mailList.append("yes")

    # update database
    for x in range(len(mailList)):
        row = x + 2
        ws[f"C{row}"] = mailList[x]

    wb.save(fnf["projectDatabase"])

    # closing smtp serever
    smtp.quit()

    thisButton.configure(text="Email Certificates")
    thisButton.configure(bg="#00ff11")
    thisButton.configure(state="normal")

    eob.configure(text=f"Done({len(data)})!")

    controller.update()


def project_settings(controller):
    # open project settings page
    controller.show_frame("projectSettings")

    # load settings data of the project
    try:
        settingsHandle = json.loads(open(fnf["projectSettings"], "r").read())
    except Exception as e:
        print("project_settings", e)
        settingsHandle = {}

    # clear the fields on the page
    emailSubjectText.delete("1.0", "end-1c")
    emailBodyText.delete("1.0", "end-1c")
    templateFileText.delete("1.0", "end-1c")

    # fill the fields on the page
    emailSubjectText.insert("end-1c", settingsHandle.get("emailSubject", ""))
    emailBodyText.insert("end-1c", settingsHandle.get("emailBody", ""))
    templateFileText.insert("end-1c", settingsHandle.get("templateFile", ""))


def project_settings_save(controller):
    """
    Save projecct settings(3)
    in project settings file
    """

    # email subject
    subjectText = emailSubjectText.get("1.0", "end-1c")
    emailSubjectText.delete("1.0", "end-1c")

    # email body
    bodyText = emailBodyText.get("1.0", "end-1c")
    emailBodyText.delete("1.0", "end-1c")

    # template file
    templateFile = templateFileText.get("1.0", "end-1c")
    templateFileText.delete("1.0", "end-1c")

    # replace new settings on settings file
    try:
        settingsHandle = json.loads(open(fnf["projectSettings"], "r").read())
    except Exception as e:
        print("Opening json", e)
        settingsHandle = {}

    settingsHandle["emailSubject"] = subjectText
    settingsHandle["emailBody"] = bodyText
    settingsHandle["templateFile"] = templateFile

    handle = open(fnf["projectSettings"], "w")
    handle.write(json.dumps(settingsHandle))
    handle.close()

    controller.show_frame("Project")


def create_records_file():
    """
    Create new Records.xlsx file
    """

    wb = oxl.Workbook()
    ws = wb.active

    # New file initiation data
    data = [
        ["Records File"],
        ["Created on:", str(datetime.datetime.now().date())],
        ["Created At:", str(datetime.datetime.now().time())],
        ["", ""],
        ["_", "_", "_", "_"],
        [" ", "  "],
        ["  ", " "],
        [" ", " "],
    ]

    for x in data:
        ws.append(x)

    # saving new file
    wb.save(fnf["baseDatabase"])


def erase_and_create_main_database(thisButton):
    """
    Delete main records excel file
    then create new file
    """
    if os.path.exists(fnf["baseDatabase"]):
        inp = messagebox.askyesno(
            "Already Exists",
            "You might already have some records in your Project directory. Overwrite?",
        )
        if inp:
            create_records_file()
            thisButton.configure(bg="#00ff11")
    else:
        create_records_file()
        thisButton.configure(bg="#00ff11")


def test(controller):
    create_records_file()
    # wb = oxl.load_workbook(fnf['projectDatabase'])
    # ws = wb.active
    # wb = oxl.load_workbook("C:\\Users\\Deekshant\\Desktop\\certi\\Certificate-Maker\\records.xlsx")
    # data = [1,2,3]
    # ws = wb.active
    # data = getData(ws)
    # ws.append(data)
    # wb.save("C:\\Users\\Deekshant\\Desktop\\certi\\Certificate-Maker\\records.xlsx")


class MainApp(tk.Tk):
    """
    App initialization class for multipage
    tkinter app loaded one on top of other
    """

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        # self.geometry("500x600")
        # self.resizable(0,0)

        # global styles
        self.title_font = tkfont.Font(family="Verdana", size=25, underline=1)
        self.buttonFont = tkfont.Font(size=15)
        self.settingsFont = tkfont.Font(family="Verdana", size=18, underline=1)

        # placement for app pages
        container = tk.Frame(self)
        container.grid(row=0, column=0)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # pages in app
        self.frames = {}
        for F in (StartPage, PageOne, Project, mainSettings, projectSettings):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        # on load page
        self.show_frame("StartPage")

    def show_frame(self, page_name):
        """
        Loads appropriate page when correct
        page name is called
        """
        frame = self.frames[page_name]
        frame.tkraise()


class StartPage(tk.Frame):
    """
    Loading page which runs as soon as the
    app opens up.
    __Certificate Maker
    - New Project
    - Existing Project
    - Database
    - Settings
    """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        # page heading
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60, pady=(10, 25))

        # New project
        button1 = tk.Button(
            self,
            text="New Project",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: new_project(controller),
        )
        button1.grid(row=1, column=0, padx=50, pady=30)

        # Existing project
        button2 = tk.Button(
            self,
            text="Existing Project",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: existing_project(controller),
        )
        button2.grid(row=2, column=0, padx=50, pady=30)

        # Database
        button3 = tk.Button(
            self,
            text="Records",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: os.startfile(f"\"{fnf['baseDatabase']}\""),
        )
        button3.grid(row=3, column=0, padx=50, pady=30)

        # Settings
        button4 = tk.Button(
            self,
            text="Settings",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: controller.show_frame("mainSettings"),
        )
        button4.grid(row=4, column=0, padx=50, pady=30)

        # button5 = tk.Button(self, text='Test', font=controller.buttonFont, padx=5, pady=2, command=lambda: test(controller))
        # button5.grid(row=5, column=0, padx=50,pady=(30,60))


class Project(tk.Frame):
    """
    Project page
    __Certificate Maker
    - Database
    - Create Certificates
    - Email certificates
    - Settings
    """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60, pady=(10, 25))

        # Database
        button1 = tk.Button(
            self,
            text="Database",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: database(button1),
        )
        button1.grid(row=1, column=0, padx=50, pady=30)

        # Certificate
        button2 = tk.Button(
            self,
            text="Create Certificates",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: create(button2, controller),
        )
        button2.grid(row=2, column=0, padx=50, pady=30)

        # Email
        button3 = tk.Button(
            self,
            text="Email Certificates",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: email(button3, controller),
        )
        button3.grid(row=3, column=0, padx=50, pady=30)

        # Settings
        button4 = tk.Button(
            self,
            text="Settings",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: project_settings(controller),
        )
        button4.grid(row=4, column=0, padx=50, pady=(30, 60))

        # Certificate generationi counter
        global aob
        aob = tk.Label(self, text="")
        aob.grid(row=2, column=0, padx=(250, 0))

        # Email sending counter
        global eob
        eob = tk.Label(self, text="")
        eob.grid(row=3, column=0, padx=(250, 0))


class mainSettings(tk.Frame):
    """
    Settings page for app
    __main settings
    - delete records
    - back
    """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        # app name
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60, pady=(10, 25))

        # main heading
        sLabel = tk.Label(self, text="Main Settings", font=controller.settingsFont)
        sLabel.grid(row=1, column=0)

        # delete records
        button1 = tk.Button(
            self,
            text="Delete Records",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: erase_and_create_main_database(button1),
        )
        button1.grid(row=2, column=0, padx=50, pady=30)

        # back-> landing page
        button4 = tk.Button(
            self,
            text="Back",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: controller.show_frame("StartPage"),
        )
        button4.grid(row=4, column=0, padx=50, pady=(30, 60))


class projectSettings(tk.Frame):
    """
    Individual settings page for all projects
    __Project Settings
    - Email Subject
    - Email Body
    - Template file
    """

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        # main heading - Certificate Maker
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.grid(row=0, column=0, padx=60, pady=(10, 15))

        # sub heading - Project Settings
        sLabel = tk.Label(self, text="Project Settings", font=controller.settingsFont)
        sLabel.grid(row=1, column=0, padx=50, pady=(0, 10))

        # Email Subject
        emailSubject = tk.Label(self, text="Email Subject :", font=("Verdana", 12))
        emailSubject.grid(row=2, column=0, sticky="W", padx=30, pady=(10, 5))

        global emailSubjectText
        emailSubjectText = tk.Text(self, bd=1, width=0, height=3, font=("Verdana", 10))
        emailSubjectText.grid(row=3, column=0, sticky="we", padx=30)

        # Email Body
        emailBody = tk.Label(self, text="Email Body :", font=("Verdana", 12))
        emailBody.grid(row=4, column=0, sticky="W", padx=30, pady=(30, 5))

        global emailBodyText
        emailBodyText = tk.Text(self, bd=1, width=0, height=5, font=("Verdana", 10))
        emailBodyText.grid(row=5, column=0, sticky="we", padx=30)

        # template file
        templateFile = tk.Label(self, text="Template File :", font=("Verdana", 12))
        templateFile.grid(row=6, column=0, sticky="W", padx=30, pady=(10, 5))

        global templateFileText
        templateFileText = tk.Text(self, bd=1, width=0, height=1, font=("Verdana", 10))
        templateFileText.grid(row=7, column=0, sticky="we", padx=30)

        # save button
        button2 = tk.Button(
            self,
            text="Save Changes",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: project_settings_save(controller),
        )
        button2.grid(row=8, column=0, sticky="W", padx=30, pady=(20, 10))

        # back button
        button4 = tk.Button(
            self,
            text="Back",
            font=controller.buttonFont,
            padx=5,
            pady=2,
            command=lambda: controller.show_frame("Project"),
        )
        button4.grid(row=8, column=0, padx=30, sticky="E", pady=(20, 0))


class PageOne(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Certificate Maker", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)
        button = tk.Button(
            self,
            text="Go to the start page",
            command=lambda: controller.show_frame("StartPage"),
        )
        button.pack()


if __name__ == "__main__":
    # basic information common to entire project
    fnf = {}

    # current location of software
    fnf["current"] = str(Path(__file__).parent.absolute())

    # current location of projects folder
    fnf["projects"] = fnf["current"] + "\\projects"

    # current loaction of gmail information file
    fnf["baseSettings"] = fnf["current"] + "\\input.dk"

    # current location of records.xlsx file
    fnf["baseDatabase"] = fnf["current"] + "\\Records.xlsx"

    # Email credentials file
    fnf["mailCreds"] = fnf["current"] + "\\creds.txt"

    # starting the app
    app = MainApp()
    app.mainloop()
